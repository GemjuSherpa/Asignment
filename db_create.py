import sqlite3
import openpyxl

#create a new db connection and cursor()
connection = sqlite3.connect("globaltemperature.db")
cursor = connection.cursor()

## creating tables>>
def create_table():
	connection.execute(
		'CREATE TABLE Bycountry(Id INTEGER PRIMARY KEY AUTOINCREMENT, '
		'Date DATE, '
		'Average_Temp REAL(4), '
		'Avg_Temp_Uncertainty REAL(4), '
		'Country VARCHAR(20));'
	)

	connection.execute(
		'CREATE TABLE Bystate(Id INTEGER PRIMARY KEY AUTOINCREMENT, '
		'Date DATE, '
		'Average_Temp REAL(4), '
		'Avg_Temp_Uncertainty REAL(4),'
		'State VARCHAR(30),'
		'Country VARCHAR(20));'
	)

	connection.execute(
		'CREATE TABLE Bycity( Id INTEGER PRIMARY KEY AUTOINCREMENT, '
		'Date    DATE, '
		'Average_Temp    REAL(4), '
		'Avg_Temp_Uncertainty    REAL(4),'
		'City   VARCHAR(30),'   
		'Country VARCHAR(30),'
		'Latitude    VARCHAR(20),'
		'Longitude   VARCHAR(20));'
	)
create_table()

## Loading all excel files and create a list of lists>>>
def load_workbook(filename, sheet):
	wb = openpyxl.load_workbook(filename)
	sheet = wb.get_sheet_by_name(sheet)
	datalists = []
	datalists.append([[i.value for i in j[0::]] for j in sheet.iter_rows(min_row=2)])

	return datalists

# Importing data into db table Bycountry..
country = load_workbook('GlobalLandTemperaturesByCountry.xlsx', 'GlobalLandTemperaturesByCountry')
for i in country:
	for j in i:
		format_str = """INSERT INTO Bycountry(Id, Date, Average_Temp, Avg_Temp_Uncertainty, Country)
		VALUES (NULL, "{Date}", "{Average_Temp}", "{Avg_Temp_Uncertainty}", "{Country}");"""
		sql_command = format_str.format(Date=j[0], Average_Temp=j[1], Avg_Temp_Uncertainty=j[2], Country=j[3])
		cursor.execute(sql_command)
		connection.commit()

## Importing data into db table Bystate
state = load_workbook('GlobalLandTemperaturesByState.xlsx', 'GlobalLandTemperaturesByState')

for i in state:
	for j in i:
		format_str = """INSERT INTO Bystate(Id, Date, Average_Temp, Avg_Temp_Uncertainty, State, Country)
		VALUES (NULL, "{Date}", "{Average_Temp}", "{Avg_Temp_Uncertainty}", "{State}", "{Country}");"""
		sql_command = format_str.format(Date=j[0], Average_Temp=j[1], Avg_Temp_Uncertainty=j[2], State=j[3], Country=j[4])
		cursor.execute(sql_command)
		connection.commit()

## Importing data into db table Bycity
city = load_workbook('GlobalLandTemperaturesByMajorCity.xlsx', 'GlobalLandTemperaturesByMajorCi')

for i in city:
	for j in i:
		format_str = """INSERT INTO Bycity(Id, Date, Average_Temp, Avg_Temp_Uncertainty, City, Country, Latitude, Longitude)
		VALUES (NULL, "{Date}", "{Average_Temp}", "{Avg_Temp_Uncertainty}", "{City}", "{Country}", "{Latitude}", "{Longitude}");"""
		sql_command = format_str.format(Date=j[0], Average_Temp=j[1], Avg_Temp_Uncertainty=j[2], City=j[3], Country=j[4], Latitude=j[5], Longitude=j[6])
		cursor.execute(sql_command)
		connection.commit()

