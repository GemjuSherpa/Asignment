# #!/bin/Python
# Filename: excel_temp.py
# Author: Gemju Sherpa

import openpyxl
import sqlite3

# Sqlite3: connect to db and query the result.
connection = sqlite3.connect("globaltemperature.db")
cursor = connection.cursor()


sql = """
	SELECT City, STRFTIME('%Y', Date), Average_Temp
	FROM Bycity
	WHERE Country = 'China'
	ORDER BY STRFTIME('%Y', Date);
"""

cursor.execute(sql)

cities = []
cities.append(cursor.fetchall())

year = []
temp = []
city = []
for lists in cities:
	for i in lists:
		year.append(i[1]) #extract the year lists
		temp.append(i[2]) #extract the temp lists
		city.append(i[0])

data = list(zip(year, temp)) # create a dictionaries k, v -> year, temp

d = {}
for k, v in data:
	d.setdefault(k, []).append(v) # combine dictionaries values together forming a lists as k, v ->year, lists of temp

values = [] #lists of Values which is temperature
keys = [] # lists if keys which is year
for k, v in d.items():
	values.append(v)
	keys.append(k)



valuesfloat = [] # lists of values converted to floats
mean = []
for j in values:
	valuesfloat.append([float(i) for i in j if i != "None"]) # eliminate the None values
for value in valuesfloat:
	mean.append(sum(value)/len(valuesfloat)) # calculate the mean temp for each year


year_mean = dict(zip(keys, mean)) # dictionaries with k,v ->year, mean_temp
city_year = dict(zip(keys, city)) # dictionaries with k,v -> year, cities

# merge a two dictionaries together so that k, v -> year, [mean, cities]
d2 = {}
for key in set(year_mean.keys()):
	try:
		d2.setdefault(key, []).append(year_mean[key])
	except KeyError:
		pass
	try:
		d2.setdefault(key, []).append(city_year[key])
	except KeyError:
		pass

## Now, create a workbook and worksheet and save the data from d2
wb = openpyxl.Workbook()
sheet = wb.create_sheet('Temperature By City', 0)

# sheet header
sheet.cell(column=1, row=1, value='Year')
sheet.cell(column=2, row=1, value='Mean_Temperature')
sheet.cell(column=3, row=1, value='City')

#iterate through the d2 keys and values
next_row=2
val = []
for key, value in d2.items():
	sheet.cell(column=1, row=next_row, value=key)
	val.append(value)
	for v in val:
		sheet.cell(column=2, row=next_row, value=v[0])
		sheet.cell(column=3, row=next_row, value=v[1])
	next_row += 1

wb.save('World Temperature.xlsx')# Save the result to workbook
wb.close() # close workbook
