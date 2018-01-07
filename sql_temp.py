import sqlite3
import openpyxl

connection = sqlite3.connect("globaltemperature.db")
cursor = connection.cursor()

def create_table():
	connection.execute(
		'CREATE TABLE Bycountry(Id INTEGER PRIMARY KEY AUTOINCREMENT, '
		'Date DATE, '
		'Average_Temp REAL(4), '
		'Avg_Temp_Uncertainty REAL(4), '
		'Country VARCHAR(20));'
	)

	connection.execute(
		'CREATE TABLE Bystate(Id INTEGER PRIMARY KEY AUTOINCREMENT, '
		'Date DATE, '
		'Average_Temp REAL(4), '
		'Avg_Temp_Uncertainty REAL(4),'
		'State VARCHAR(30),'
		'Country VARCHAR(20));'
	)

	connection.execute(
		'CREATE TABLE Bycity( Id INTEGER PRIMARY KEY AUTOINCREMENT, '
		'Date    DATE, '
		'Average_Temp    REAL(4), '
		'Avg_Temp_Uncertainty    REAL(4),'
		'City   VARCHAR(30),'   
		'Country VARCHAR(30),'
		'Latitude    VARCHAR(20),'
		'Longitude   VARCHAR(20));'
	)
#create_table()

def load_workbook(filename, sheet):
	wb = openpyxl.load_workbook(filename)
	sheet = wb.get_sheet_by_name(sheet)
	datalists = []
	datalists.append([[i.value for i in j[0::]] for j in sheet.rows])

	return datalists

country = load_workbook('GlobalLandTemperaturesByCountry.xlsx', 'GlobalLandTemperaturesByCountry')
city = load_workbook('GlobalLandTemperaturesByState.xlsx', 'GlobalLandTemperaturesByState')
state = load_workbook('GlobalLandTemperaturesByMajorCity.xlsx', 'GlobalLandTemperaturesByMajorCity')
for i in country:
	for j in i:
		format_str = """INSERT INTO Bycountry(Id, Date, Average_Temp, Avg_Temp_Uncertainty, Country)
		VALUES (NULL, "{Date}", "{Average_Temp}", "{Avg_Temp_Uncertainty}", "{Country}");"""
		sql_command = format_str.format(Date=j[0], Average_Temp=j[1], Avg_Temp_Uncertainty=j[2], Country=j[3])
		cursor.execute(sql_command)
		connection.commit()


